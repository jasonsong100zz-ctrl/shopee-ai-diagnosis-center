#!/usr/bin/env python3
"""Prepare, annotate, aggregate, render, and verify ecommerce review decision-driver runs."""

from __future__ import annotations

import argparse
import csv
import hashlib
import ipaddress
import itertools
import json
import re
import socket
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Iterator


SKILL_ROOT = Path(__file__).resolve().parents[1]
TAXONOMY_PATH = SKILL_ROOT / "assets" / "factor-taxonomy.json"
TEMPLATE_PATH = SKILL_ROOT / "assets" / "dashboard_template.html"
MAX_IMAGE_BYTES = 8 * 1024 * 1024
CONFIRMED_STATUSES = {"ai_confirmed", "human_confirmed"}
CONFIRMED_TRANSLATIONS = {"original_zh", "ai_confirmed", "human_confirmed"}


FIELD_ALIASES = {
    "product_id": ["product id", "productid", "item id", "itemid", "asin", "sku", "商品id", "产品id", "商品编号", "产品编号", "子asin", "商品asin", "产品asin", "id produk", "kode produk", "รหัสสินค้า"],
    "brand": ["brand", "品牌", "merek", "merk", "แบรนด์"],
    "product_name": ["product name", "product title", "item name", "商品名称", "产品名称", "商品标题", "产品标题", "nama produk", "ชื่อสินค้า"],
    "product_category": ["category", "product category", "品类", "类目", "产品类别", "商品类别", "kategori", "หมวดหมู่"],
    "product_image_urls": ["product image", "product images", "main image", "main image url", "产品主图", "商品主图", "产品图片", "商品图片", "主图链接", "foto produk", "gambar produk", "รูปสินค้า"],
    "platform": ["platform", "渠道", "平台", "source platform", "marketplace", "แพลตฟอร์ม"],
    "market": ["market", "country", "site", "站点", "国家", "市场", "negara", "ประเทศ"],
    "language": ["language", "语言", "语种", "bahasa", "ภาษา"],
    "entity_role": ["entity role", "product role", "role", "产品角色", "对象角色", "peran produk"],
    "review_id": ["review id", "reviewid", "comment id", "feedback id", "评论id", "评价id", "id ulasan", "รหัสรีวิว"],
    "review_title": ["review title", "headline", "summary", "评论标题", "评价标题", "judul ulasan", "หัวข้อรีวิว"],
    "review_body": ["review body", "review text", "review content", "comment", "content", "body", "text", "评论内容", "评价内容", "评论正文", "评价正文", "内容", "正文", "ulasan", "isi ulasan", "komentar", "รีวิว", "ความคิดเห็น"],
    "rating": ["rating", "stars", "star", "score", "星级", "评分", "打分", "评价等级", "bintang", "คะแนน", "ดาว"],
    "review_date": ["review date", "created at", "date", "time", "评论日期", "评价日期", "评论时间", "评价时间", "tanggal ulasan", "วันที่รีวิว"],
    "variant": ["variant", "variation", "style", "size", "shade", "color", "变体", "规格", "色号", "款式", "varian", "รุ่น", "ขนาด"],
    "verified": ["verified purchase", "verified", "已验证购买", "pembelian terverifikasi", "ยืนยันการซื้อ"],
    "helpful_votes": ["helpful votes", "helpful", "likes", "有用数", "点赞数", "membantu", "มีประโยชน์"],
    "review_image_urls": ["review images", "review image urls", "media urls", "images", "评论图片", "评价图片", "买家秀", "图片链接", "gambar ulasan", "foto ulasan", "รูปรีวิว"],
}

POSITIVE_HINTS = [
    "喜欢", "满意", "很好", "好用", "推荐", "回购", "有效", "舒服",
    "love", "great", "good", "perfect", "recommend", "satisfied", "works", "comfortable", "easy to use", "lightweight",
    "bagus", "suka", "cocok", "nyaman", "berhasil", "rekomen", "beli lagi", "cepat meresap", "mudah",
    "ดี", "ชอบ", "แนะนำ", "ได้ผล", "สบาย", "ซื้อซ้ำ", "หอม", "ง่าย", "方便", "清爽",
]
NEGATIVE_HINTS = [
    "不喜欢", "失望", "糟糕", "坏了", "破损", "退货", "退款", "不会再买", "无效", "过敏",
    "disappointed", "bad", "broken", "return", "refund", "never again", "doesn't", "does not", "didn't", "poor", "irritation",
    "tidak", "nggak", "gak", "kecewa", "rusak", "mahal", "alergi", "iritasi",
    "ไม่", "ไม่เหมาะ", "ผิดหวัง", "พัง", "แพ้", "ระคายเคือง", "แพง", "คืนเงิน", "价格偏贵", "容量少", "容量有点少",
]
IMPACT_KEYWORDS = {
    "RETURN_REFUND": ["退货", "退款", "return", "refund", "kembalikan", "pengembalian dana", "คืนสินค้า", "คืนเงิน"],
    "STOP_USE": ["停止使用", "不敢再用", "stopped using", "stop using", "berhenti pakai", "หยุดใช้"],
    "SAFETY": ["过敏", "红肿", "刺痛", "灼烧", "起痘", "allergy", "rash", "stinging", "burning", "breakout", "alergi", "iritasi", "panas", "แพ้", "ระคายเคือง", "แสบ"],
    "UNUSABLE": ["无法使用", "完全不能用", "unusable", "cannot use", "can't use", "tidak bisa dipakai", "ใช้ไม่ได้"],
    "NO_REPURCHASE": ["不会回购", "不会再买", "never repurchase", "won't buy again", "tidak beli lagi", "ไม่ซื้อซ้ำ"],
    "SWITCH_COMPETITOR": ["换别的品牌", "买了竞品", "switching to", "another brand", "ganti merek", "เปลี่ยนแบรนด์"],
    "REPURCHASE": ["会回购", "已经回购", "will repurchase", "bought again", "beli lagi", "ซื้อซ้ำ"],
    "RECOMMEND": ["推荐", "recommend", "rekomen", "แนะนำ"],
}

GENERIC_KEYWORDS = {"good", "great", "bad", "product", "item", "love", "bagus", "produk", "很好", "产品", "商品", "ดี", "สินค้า"}


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    result = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            value = json.loads(line)
            if not isinstance(value, dict):
                raise ValueError(f"{path}:{line_number} 必须是 JSON 对象")
            result.append(value)
    return result


def stable_hash(*values: Any, length: int = 16) -> str:
    payload = "\u241f".join(str(value or "").strip() for value in values)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:length]


def text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"[\W_]+", "", text_value(value).casefold(), flags=re.UNICODE)


NORMALIZED_ALIASES = {field: [normalize_header(alias) for alias in aliases] for field, aliases in FIELD_ALIASES.items()}


def match_field(header: Any) -> tuple[str | None, int]:
    normalized = normalize_header(header)
    best_field, best_score = None, 0
    for field, aliases in NORMALIZED_ALIASES.items():
        for alias in aliases:
            if normalized == alias:
                score = 1000 + len(alias)
            elif len(alias) >= 5 and (normalized.startswith(alias) or normalized.endswith(alias)):
                score = 500 + len(alias)
            elif len(alias) >= 6 and alias in normalized:
                score = 250 + len(alias)
            else:
                continue
            if score > best_score:
                best_field, best_score = field, score
    return best_field, best_score


def map_headers(headers: list[Any]) -> dict[str, int]:
    candidates: dict[str, tuple[int, int]] = {}
    for index, header in enumerate(headers):
        field, score = match_field(header)
        if field and (field not in candidates or score > candidates[field][1]):
            candidates[field] = (index, score)
    return {field: item[0] for field, item in candidates.items()}


def find_header_row(rows: list[list[Any]], limit: int = 20) -> tuple[int, dict[str, int]]:
    best = (-1, {}, -1)
    for index, row in enumerate(rows[:limit]):
        mapping = map_headers(list(row))
        score = len(mapping) * 10 + int("review_body" in mapping) * 30 + int("product_id" in mapping) * 12
        if score > best[2]:
            best = (index, mapping, score)
    if "review_body" not in best[1] and "review_title" not in best[1]:
        raise ValueError("找不到评论正文或评论标题列")
    return best[0], best[1]


def iter_xlsx(path: Path) -> Iterator[tuple[str, int, dict[str, Any], dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 XLSX 需要 openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            preview = [list(row) for row in itertools.islice(iterator, 30)]
            if not preview:
                continue
            try:
                header_index, mapping = find_header_row(preview)
            except ValueError:
                continue
            headers = [text_value(value) for value in preview[header_index]]
            labels = {field: headers[index] for field, index in mapping.items()}
            rows = itertools.chain(preview[header_index + 1 :], iterator)
            for row_number, row in enumerate(rows, start=header_index + 2):
                values = list(row)
                if not any(text_value(value) for value in values):
                    continue
                yield sheet.title, row_number, {field: values[index] if index < len(values) else None for field, index in mapping.items()}, labels
    finally:
        workbook.close()


def decode_delimited(path: Path) -> str:
    payload = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "windows-1252"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="replace")


def iter_delimited(path: Path) -> Iterator[tuple[str, int, dict[str, Any], dict[str, str]]]:
    text = decode_delimited(path)
    delimiter = "\t" if path.suffix.casefold() == ".tsv" else ","
    try:
        delimiter = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|").delimiter
    except csv.Error:
        pass
    rows = [list(row) for row in csv.reader(text.splitlines(), delimiter=delimiter)]
    header_index, mapping = find_header_row(rows)
    headers = [text_value(value) for value in rows[header_index]]
    labels = {field: headers[index] for field, index in mapping.items()}
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if any(text_value(value) for value in row):
            yield "data", row_number, {field: row[index] if index < len(row) else None for field, index in mapping.items()}, labels


def iter_json(path: Path) -> Iterator[tuple[str, int, dict[str, Any], dict[str, str]]]:
    value = read_json(path)
    if isinstance(value, dict):
        value = value.get("reviews", value.get("data", value.get("rows")))
    if not isinstance(value, list):
        raise ValueError("JSON 必须是对象数组，或包含 reviews/data/rows 数组")
    headers = sorted({str(key) for row in value if isinstance(row, dict) for key in row})
    mapping = map_headers(headers)
    if "review_body" not in mapping and "review_title" not in mapping:
        raise ValueError("JSON 找不到评论正文字段")
    labels = {field: headers[index] for field, index in mapping.items()}
    for row_number, row in enumerate(value, start=1):
        if isinstance(row, dict):
            yield "json", row_number, {field: row.get(headers[index]) for field, index in mapping.items()}, labels


def iter_source(path: Path) -> Iterator[tuple[str, int, dict[str, Any], dict[str, str]]]:
    if path.suffix.casefold() == ".xlsx":
        yield from iter_xlsx(path)
    elif path.suffix.casefold() in {".csv", ".tsv"}:
        yield from iter_delimited(path)
    elif path.suffix.casefold() == ".json":
        yield from iter_json(path)
    else:
        raise ValueError(f"不支持的文件格式：{path.suffix}")


def parse_rating(value: Any) -> float | None:
    match = re.search(r"([1-5](?:\.\d+)?)", text_value(value))
    if not match:
        return None
    rating = float(match.group(1))
    return rating if 0 < rating <= 5 else None


def parse_int(value: Any) -> int:
    match = re.search(r"-?\d+", text_value(value).replace(",", ""))
    return int(match.group(0)) if match else 0


def parse_bool(value: Any) -> bool | None:
    normalized = text_value(value).casefold()
    if normalized in {"true", "yes", "y", "1", "是", "verified", "terverifikasi", "ยืนยัน"}:
        return True
    if normalized in {"false", "no", "n", "0", "否", "unverified", "tidak", "ไม่"}:
        return False
    return None


URL_PATTERN = re.compile(r"https?://[^\s,;|\]\[\"']+", re.IGNORECASE)


def parse_urls(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        values = [text_value(item) for item in value]
    else:
        raw = text_value(value)
        if not raw:
            return []
        try:
            decoded = json.loads(raw)
            values = [text_value(item) for item in decoded] if isinstance(decoded, list) else [raw]
        except json.JSONDecodeError:
            values = [raw]
    urls: list[str] = []
    for item in values:
        matches = URL_PATTERN.findall(item)
        urls.extend(matches or [part.strip() for part in re.split(r"[,;|\n]+", item) if part.strip().startswith(("http://", "https://"))])
    return list(dict.fromkeys(url.rstrip(".)") for url in urls))


def normalize_role(value: Any) -> str:
    raw = text_value(value).casefold()
    if raw in {"target", "self", "own", "目标", "自有", "本品", "produk sendiri", "สินค้าเป้าหมาย"}:
        return "TARGET"
    return "PEER"


def infer_source_polarity(source_polarity: str, rating: float | None) -> str:
    if source_polarity in {"positive", "negative"}:
        return source_polarity
    if rating is None:
        return "unknown"
    if rating >= 4:
        return "positive"
    if rating <= 2:
        return "negative"
    return "neutral"


def infer_product_id_from_filename(path: Path) -> str:
    patterns = [
        r"(?:产品|商品)\s*id\s*[=：:]\s*([A-Za-z0-9_-]+)",
        r"product\s*id\s*[=：:]\s*([A-Za-z0-9_-]+)",
        r"item\s*id\s*[=：:]\s*([A-Za-z0-9_-]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, path.stem, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return ""


def infer_market_from_filename(path: Path) -> str:
    name = path.stem.casefold()
    aliases = {
        "印尼": "ID",
        "indonesia": "ID",
        "泰国": "TH",
        "thailand": "TH",
        "美国": "US",
        "united states": "US",
        "马来西亚": "MY",
        "malaysia": "MY",
        "菲律宾": "PH",
        "philippines": "PH",
        "越南": "VN",
        "vietnam": "VN",
    }
    return next((market for marker, market in aliases.items() if marker in name), "")


def infer_platform_from_urls(urls: list[str]) -> str:
    hosts = " ".join(urllib.parse.urlparse(url).hostname or "" for url in urls).casefold()
    if "shopee" in hosts or "susercontent.com" in hosts:
        return "Shopee"
    if "amazon" in hosts or "media-amazon.com" in hosts:
        return "Amazon"
    if "tiktok" in hosts or "byteoversea.com" in hosts:
        return "TikTok Shop"
    return ""


def detect_language(text: str, supplied: str = "") -> str:
    supplied = supplied.strip().casefold().replace("_", "-")
    aliases = {"chinese": "zh", "中文": "zh", "english": "en", "英文": "en", "indonesian": "id", "bahasa indonesia": "id", "印尼语": "id", "thai": "th", "泰语": "th"}
    if supplied:
        return aliases.get(supplied, supplied.split("-")[0])
    if re.search(r"[\u0e00-\u0e7f]", text):
        return "th"
    if re.search(r"[\u3400-\u9fff]", text):
        return "zh"
    lowered = f" {text.casefold()} "
    id_markers = [" yang ", " tidak ", " nggak ", " gak ", " sangat ", " cocok ", " kulit ", " lembap", " bagus", " produk ", " tapi ", " untuk ", " cepat ", " bikin "]
    if sum(marker in lowered for marker in id_markers) >= 1:
        return "id"
    return "en" if re.search(r"[a-z]", lowered) else "und"


def normalize_record(path: Path, entity_role: str, source_polarity: str, sheet: str, row_number: int, record: dict[str, Any], category_override: str, target_product_id: str) -> dict[str, Any] | None:
    body = text_value(record.get("review_body"))
    title = text_value(record.get("review_title"))
    raw_text = ". ".join(part for part in (title, body) if part)
    if not raw_text:
        return None
    product_name = text_value(record.get("product_name"))
    product_id = re.sub(r"\s+", "-", (text_value(record.get("product_id")) or infer_product_id_from_filename(path)).strip())
    temporary = not bool(product_id)
    if temporary:
        product_id = f"UNKNOWN-{stable_hash(text_value(record.get('platform')), text_value(record.get('market')), product_name or path.name, length=10).upper()}"
    role = entity_role if entity_role in {"TARGET", "PEER"} else normalize_role(record.get("entity_role"))
    if target_product_id and product_id.casefold() == target_product_id.casefold():
        role = "TARGET"
    rating = parse_rating(record.get("rating"))
    product_images = parse_urls(record.get("product_image_urls"))
    review_images = parse_urls(record.get("review_image_urls"))
    platform = text_value(record.get("platform")) or infer_platform_from_urls(product_images + review_images) or "UNKNOWN"
    market = text_value(record.get("market")).upper() or infer_market_from_filename(path) or "UNKNOWN"
    source_review_id = text_value(record.get("review_id"))
    review_id = f"review-{stable_hash(platform, market, product_id, source_review_id or rating, record.get('review_date'), raw_text)}"
    language = detect_language(raw_text, text_value(record.get("language")))
    return {
        "review_id": review_id,
        "source_review_id": source_review_id,
        "product_id": product_id,
        "temporary_product_id": temporary,
        "entity_role": role,
        "brand": text_value(record.get("brand")) or "未知品牌",
        "product_name": product_name or product_id,
        "product_category": category_override or text_value(record.get("product_category")),
        "variant": text_value(record.get("variant")),
        "platform": platform,
        "market": market,
        "language": language,
        "product_images": [{"url": url, "src": url, "cached": False} for url in product_images],
        "review_title": title,
        "review_body": body,
        "raw_text": raw_text,
        "rating": rating,
        "review_date": text_value(record.get("review_date")),
        "verified": parse_bool(record.get("verified")),
        "helpful_votes": parse_int(record.get("helpful_votes")),
        "review_images": [{"url": url, "src": url, "cached": False} for url in review_images],
        "source_polarity": infer_source_polarity(source_polarity, rating),
        "source_file": str(path.resolve()),
        "source_sheet": sheet,
        "source_row": row_number,
    }


def review_dedupe_key(review: dict[str, Any]) -> str:
    if review.get("source_review_id"):
        return stable_hash(review["platform"], review["market"], review["product_id"], review["source_review_id"])
    body = re.sub(r"\s+", " ", review["raw_text"].casefold()).strip()
    return stable_hash(review["platform"], review["market"], review["product_id"], review["rating"], review["review_date"], body)


def cross_product_text_clusters(reviews: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for review in reviews:
        body = re.sub(r"\s+", " ", review["raw_text"].casefold()).strip()
        grouped[stable_hash(review["language"], body, length=64)].append(review)
    rows = []
    for digest, values in grouped.items():
        products = sorted({value["product_id"] for value in values})
        if len(products) > 1:
            rows.append({"text_hash": digest, "row_count": len(values), "product_ids": products, "review_ids": [value["review_id"] for value in values], "evidence_excerpt": values[0]["raw_text"][:240]})
    return sorted(rows, key=lambda row: (-row["row_count"], row["text_hash"]))


def safe_hostname(hostname: str) -> bool:
    if not hostname or hostname.casefold() in {"localhost", "localhost.localdomain"}:
        return False
    try:
        infos = socket.getaddrinfo(hostname, None, type=socket.SOCK_STREAM)
    except socket.gaierror:
        return False
    return all(not any((address.is_private, address.is_loopback, address.is_link_local, address.is_reserved, address.is_multicast)) for address in (ipaddress.ip_address(info[4][0]) for info in infos))


def validate_remote_url(url: str) -> None:
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in {"http", "https"} or not safe_hostname(parsed.hostname or ""):
        raise ValueError("图片 URL 不是允许的公网 HTTP(S) 地址")


class SafeRedirectHandler(urllib.request.HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):  # type: ignore[no-untyped-def]
        validate_remote_url(newurl)
        return super().redirect_request(req, fp, code, msg, headers, newurl)


def download_image(url: str, output_dir: Path, prefix: str) -> tuple[str, str | None]:
    try:
        validate_remote_url(url)
        request = urllib.request.Request(url, headers={"User-Agent": "ecommerce-review-purchase-drivers/1.0"})
        with urllib.request.build_opener(SafeRedirectHandler()).open(request, timeout=12) as response:
            content_type = (response.headers.get_content_type() or "").casefold()
            if not content_type.startswith("image/"):
                raise ValueError("响应不是图片")
            payload = response.read(MAX_IMAGE_BYTES + 1)
            if len(payload) > MAX_IMAGE_BYTES:
                raise ValueError("图片超过 8 MB")
        extension = {"image/jpeg": ".jpg", "image/png": ".png", "image/webp": ".webp", "image/gif": ".gif", "image/avif": ".avif"}.get(content_type, ".img")
        output_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{prefix}-{stable_hash(url, length=20)}{extension}"
        (output_dir / filename).write_bytes(payload)
        return filename, None
    except (ValueError, OSError, urllib.error.URLError) as exc:
        return "", str(exc)


def cache_images(reviews: list[dict[str, Any]], output_dir: Path, enabled: bool) -> dict[str, Any]:
    domains, failures, cached = Counter(), [], 0
    memo: dict[tuple[str, str], tuple[str, str | None]] = {}
    for review in reviews:
        for kind, field in (("product", "product_images"), ("review", "review_images")):
            for image in review[field]:
                host = urllib.parse.urlparse(image["url"]).hostname
                if host:
                    domains[host.casefold()] += 1
                if not enabled:
                    continue
                key = (kind, image["url"])
                if key not in memo:
                    relative_dir = Path("assets") / f"{kind}-images"
                    filename, error = download_image(image["url"], output_dir / relative_dir, kind)
                    memo[key] = (str(relative_dir / filename) if filename else "", error)
                relative, error = memo[key]
                if relative:
                    image.update({"src": relative, "cached": True})
                    cached += 1
                else:
                    image["error"] = error
                    failures.append({"url": image["url"], "error": error or "unknown"})
    return {"download_enabled": enabled, "domains": dict(domains.most_common()), "cached_count": cached, "failure_count": len(failures), "failures": failures[:100]}


ATOM_SPLIT = re.compile(
    r"(?<=[。！？!?；;\.])(?:\s+|(?=[A-Z-]))|[。！？!?；;,，\n]+|"
    r"(?=\b(?:but|however|although|yet|tapi|namun|and|dan)\b)|"
    r"(?=(?:但是|不过|然而|可是|但(?!是)|และ|แต่))",
    re.IGNORECASE,
)


def split_atoms(text: str) -> list[tuple[str, int, int]]:
    parts = [part.strip(" ，,。.!！?？;；") for part in ATOM_SPLIT.split(text) if part and part.strip(" ，,。.!！?？;；")]
    if not parts and text.strip():
        parts = [text.strip()]
    atoms, cursor = [], 0
    for part in parts:
        start = text.find(part, cursor)
        if start < 0:
            start = max(0, cursor)
        atoms.append((part, start, start + len(part)))
        cursor = start + len(part)
    return atoms


def infer_polarity(text: str, source_polarity: str) -> str:
    lowered = text.casefold()
    positive = sum(1 for hint in POSITIVE_HINTS if hint in lowered)
    negative = sum(1 for hint in NEGATIVE_HINTS if hint in lowered)
    negated_problem = re.search(
        r"(?:does not|doesn't|not|tidak|nggak|gak|ไม่|不|没有)\s*"
        r"(?:leak|leaking|greasy|sticky|irritat|bocor|berminyak|lengket|เหนียว|มัน|รั่ว|油腻|粘|漏|刺痛)",
        lowered,
    )
    if negated_problem and negative <= positive + 1:
        return "positive"
    if negative > positive:
        return "negative"
    if positive > negative:
        return "positive"
    return source_polarity if source_polarity in {"positive", "negative", "neutral"} else "neutral"


def infer_impacts(text: str) -> list[str]:
    lowered = text.casefold()
    result = [signal for signal, keywords in IMPACT_KEYWORDS.items() if any(keyword.casefold() in lowered for keyword in keywords)]
    negated_reaction = re.search(r"(?:no|not|does not|doesn't|tidak|nggak|gak|ไม่|不|没有)\s*(?:allerg|rash|irritat|sting|burn|breakout|alergi|iritasi|แพ้|ระคายเคือง|แสบ|过敏|红肿|刺痛|灼烧|起痘)", lowered)
    if negated_reaction and "SAFETY" in result:
        result.remove("SAFETY")
    if "NO_REPURCHASE" in result and "REPURCHASE" in result:
        result.remove("REPURCHASE")
    return result


def candidate_factor(text: str, taxonomy: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    lowered = text.casefold()
    scored = []
    for factor in taxonomy["seed_factors"]:
        matches = [alias for alias in factor.get("aliases", []) if alias.casefold() in lowered]
        score = sum(max(1, len(alias)) for alias in matches)
        if score:
            scored.append((score, factor, matches))
    if not scored:
        fallback = next(row for row in taxonomy["seed_factors"] if row["id"] == "insufficient_detail")
        return fallback, []
    scored.sort(key=lambda row: (-row[0], row[1]["id"]))
    return scored[0][1], scored[0][2]


def build_atoms(reviews: list[dict[str, Any]], taxonomy: dict[str, Any]) -> list[dict[str, Any]]:
    atoms = []
    for review in reviews:
        for ordinal, (evidence, start, end) in enumerate(split_atoms(review["raw_text"]), start=1):
            language = detect_language(evidence, review["language"] if review["language"] not in {"und", ""} else "")
            factor, matches = candidate_factor(evidence, taxonomy)
            polarity = infer_polarity(evidence, review["source_polarity"])
            atoms.append({
                "atom_id": f"atom-{stable_hash(review['platform'], review['market'], review['product_id'], review['review_id'], ordinal, evidence)}",
                "review_id": review["review_id"],
                "product_id": review["product_id"],
                "entity_role": review["entity_role"],
                "brand": review["brand"],
                "platform": review["platform"],
                "market": review["market"],
                "ordinal": ordinal,
                "char_start": start,
                "char_end": end,
                "evidence_original": evidence,
                "evidence_zh": evidence if language == "zh" else "",
                "language": language,
                "polarity": polarity,
                "factor_id": factor["id"],
                "family_id": factor["family_id"],
                "keywords_original": matches,
                "keywords_canonical_zh": [factor["label_zh"]] if factor["id"] != "insufficient_detail" else [],
                "contexts": [],
                "impact_signals": infer_impacts(evidence),
                "confidence": 0.35 if matches else 0.15,
                "translation_status": "original_zh" if language == "zh" else "pending",
                "review_status": "candidate_rule",
            })
    return atoms


def build_factor_catalog_draft(atoms: list[dict[str, Any]], taxonomy: dict[str, Any], category: str) -> dict[str, Any]:
    used = Counter(atom["factor_id"] for atom in atoms)
    factors = []
    for factor in taxonomy["seed_factors"]:
        if used[factor["id"]] or factor["id"] == "insufficient_detail":
            factors.append({"id": factor["id"], "family_id": factor["family_id"], "label_zh": factor["label_zh"], "definition": "由通用种子词提出，需结合本品类样本确认边界。", "sample_atom_count": used[factor["id"]]})
    return {"schema_version": "decision-factor-catalog-v1", "category": category, "status": "candidate_rule", "factors": factors}


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            serialized = {key: json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value for key, value in row.items()}
            writer.writerow(serialized)


def command_prepare(args: argparse.Namespace) -> int:
    taxonomy = read_json(TAXONOMY_PATH)
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    specs: list[tuple[Path, str, str]] = []
    specs += [(Path(value).resolve(), "AUTO", "auto") for value in args.input]
    specs += [(Path(value).resolve(), "TARGET", "auto") for value in args.target]
    specs += [(Path(value).resolve(), "PEER", "auto") for value in args.peer]
    specs += [(Path(value).resolve(), "AUTO", "positive") for value in args.positive]
    specs += [(Path(value).resolve(), "AUTO", "negative") for value in args.negative]
    if not specs:
        raise ValueError("至少提供一个 --input、--target、--peer、--positive 或 --negative 文件")
    reviews, duplicates, source_audit = [], 0, []
    seen = set()
    for path, role, polarity in specs:
        if not path.is_file():
            raise FileNotFoundError(path)
        count, mappings = 0, []
        for sheet, row_number, record, labels in iter_source(path):
            normalized = normalize_record(path, role, polarity, sheet, row_number, record, args.category or "", args.target_product_id or "")
            if not normalized:
                continue
            key = review_dedupe_key(normalized)
            if key in seen:
                duplicates += 1
                continue
            seen.add(key)
            reviews.append(normalized)
            count += 1
            mappings.append(labels)
        source_audit.append({"file": str(path), "entity_role_override": role, "polarity_override": polarity, "imported_review_count": count, "detected_columns": mappings[0] if mappings else {}})
    if not reviews:
        raise ValueError("输入文件中没有有效评论")
    image_audit = cache_images(reviews, output_dir, args.download_images)
    clusters = cross_product_text_clusters(reviews)
    atoms = build_atoms(reviews, taxonomy)
    categories = sorted({review["product_category"] for review in reviews if review["product_category"]})
    homogeneity = "CONFIRMED" if len(categories) == 1 else "MIXED" if len(categories) > 1 else "UNCONFIRMED"
    category = categories[0] if len(categories) == 1 else args.category or ""
    draft = build_factor_catalog_draft(atoms, taxonomy, category)
    review_by_id = {review["review_id"]: review for review in reviews}
    queue = []
    for atom in atoms:
        review = review_by_id[atom["review_id"]]
        queue.append({**atom, "product_name": review["product_name"], "product_category": review["product_category"], "rating": review["rating"], "review_date": review["review_date"], "raw_review": review["raw_text"], "source_file": review["source_file"], "source_sheet": review["source_sheet"], "source_row": review["source_row"]})
    discovery = sorted(queue, key=lambda row: (row["language"], row["product_id"], row["polarity"], row["atom_id"]))[:500]
    write_jsonl(output_dir / "normalized-reviews.jsonl", reviews)
    write_csv(output_dir / "normalized-reviews.csv", reviews, ["review_id", "source_review_id", "product_id", "entity_role", "brand", "product_name", "product_category", "variant", "platform", "market", "language", "rating", "review_date", "verified", "helpful_votes", "source_polarity", "raw_text", "product_images", "review_images", "source_file", "source_sheet", "source_row"])
    write_jsonl(output_dir / "evidence-atoms.jsonl", atoms)
    write_jsonl(output_dir / "factor-discovery-queue.jsonl", discovery)
    write_jsonl(output_dir / "annotation-queue.jsonl", queue)
    write_json(output_dir / "factor-catalog-draft.json", draft)
    audit = {
        "schema_version": "review-purchase-driver-import-v1",
        "prepared_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "demo_mode": bool(args.demo),
        "review_count": len(reviews),
        "atom_count": len(atoms),
        "product_count": len({review["product_id"] for review in reviews}),
        "brand_count": len({review["brand"] for review in reviews}),
        "target_product_count": len({review["product_id"] for review in reviews if review["entity_role"] == "TARGET"}),
        "temporary_product_review_count": sum(review["temporary_product_id"] for review in reviews),
        "duplicate_count": duplicates,
        "category_values": categories,
        "category_homogeneity": homogeneity,
        "language_counts": dict(Counter(review["language"] for review in reviews)),
        "market_counts": dict(Counter(review["market"] for review in reviews)),
        "platform_counts": dict(Counter(review["platform"] for review in reviews)),
        "cross_product_repeated_text_cluster_count": len(clusters),
        "cross_product_repeated_text_row_count": sum(row["row_count"] for row in clusters),
        "cross_product_repeated_text_clusters": clusters[:500],
        "review_with_images_count": sum(bool(review["review_images"]) for review in reviews),
        "image_audit": image_audit,
        "sources": source_audit,
    }
    write_json(output_dir / "import-audit.json", audit)
    print(json.dumps({"status": "PREPARED", "output_dir": str(output_dir), "reviews": len(reviews), "products": audit["product_count"], "languages": audit["language_counts"], "atoms": len(atoms), "next": ["确认 factor-catalog-draft.json 并保存为 factor-catalog.json", "完成 annotation-queue.jsonl 后运行 build"]}, ensure_ascii=False))
    return 0


def validate_catalog(catalog: dict[str, Any], taxonomy: dict[str, Any]) -> tuple[dict[str, dict[str, Any]], bool]:
    family_ids = {row["id"] for row in taxonomy["families"]}
    factors, errors = {}, []
    for index, factor in enumerate(catalog.get("factors", []), start=1):
        factor_id = text_value(factor.get("id"))
        family_id = text_value(factor.get("family_id"))
        label = text_value(factor.get("label_zh"))
        if not factor_id or factor_id in factors:
            errors.append(f"因子目录第 {index} 行 ID 缺失或重复")
        elif family_id not in family_ids:
            errors.append(f"因子 {factor_id} 的 family_id 无效")
        elif not label:
            errors.append(f"因子 {factor_id} 缺少 label_zh")
        else:
            factors[factor_id] = {**factor, "id": factor_id, "family_id": family_id, "label_zh": label}
    if errors:
        raise ValueError("因子目录校验失败：\n- " + "\n- ".join(errors[:50]))
    return factors, catalog.get("status") in CONFIRMED_STATUSES


def validate_annotations(annotations: list[dict[str, Any]], atoms: list[dict[str, Any]], factors: dict[str, dict[str, Any]], taxonomy: dict[str, Any]) -> dict[str, dict[str, Any]]:
    atom_ids = {atom["atom_id"] for atom in atoms}
    impacts = set(taxonomy["impact_signals"])
    result, errors = {}, []
    for index, row in enumerate(annotations, start=1):
        atom_id = text_value(row.get("atom_id"))
        factor_id = text_value(row.get("factor_id"))
        polarity = text_value(row.get("polarity"))
        status = text_value(row.get("review_status"))
        translation_status = text_value(row.get("translation_status"))
        if atom_id not in atom_ids or atom_id in result:
            errors.append(f"标注第 {index} 行 atom_id 无效或重复")
            continue
        if factor_id not in factors:
            errors.append(f"{atom_id} 的 factor_id 不在运行目录中")
        if polarity not in {"positive", "negative", "neutral"}:
            errors.append(f"{atom_id} 的 polarity 无效")
        if status not in CONFIRMED_STATUSES | {"candidate_rule", "needs_review"}:
            errors.append(f"{atom_id} 的 review_status 无效")
        if translation_status not in CONFIRMED_TRANSLATIONS | {"pending", "needs_review"}:
            errors.append(f"{atom_id} 的 translation_status 无效")
        try:
            confidence = float(row.get("confidence"))
        except (TypeError, ValueError):
            confidence = -1
        if not 0 <= confidence <= 1:
            errors.append(f"{atom_id} 的 confidence 必须在 0–1")
        row_impacts = row.get("impact_signals") or []
        if not isinstance(row_impacts, list) or any(value not in impacts for value in row_impacts):
            errors.append(f"{atom_id} 包含未知 impact_signals")
        result[atom_id] = {
            "factor_id": factor_id,
            "polarity": polarity,
            "language": text_value(row.get("language")),
            "evidence_zh": text_value(row.get("evidence_zh")),
            "keywords_original": [text_value(value) for value in row.get("keywords_original", []) if text_value(value)],
            "keywords_canonical_zh": [text_value(value) for value in row.get("keywords_canonical_zh", []) if text_value(value)],
            "contexts": [text_value(value) for value in row.get("contexts", []) if text_value(value)],
            "impact_signals": list(dict.fromkeys(row_impacts)) if isinstance(row_impacts, list) else [],
            "confidence": round(confidence, 3),
            "translation_status": translation_status,
            "review_status": status,
        }
    if errors:
        raise ValueError("标注校验失败：\n- " + "\n- ".join(errors[:50]))
    return result


def enrich_atoms(atoms: list[dict[str, Any]], annotations: dict[str, dict[str, Any]], factors: dict[str, dict[str, Any]], taxonomy: dict[str, Any]) -> list[dict[str, Any]]:
    families = {row["id"]: row for row in taxonomy["families"]}
    enriched = []
    for source in atoms:
        atom = dict(source)
        if atom["atom_id"] in annotations:
            atom.update(annotations[atom["atom_id"]])
        factor = factors.get(atom.get("factor_id"))
        atom["factor_label"] = factor["label_zh"] if factor else "未分类"
        atom["family_id"] = factor["family_id"] if factor else ""
        atom["family_label"] = families.get(atom["family_id"], {}).get("label_zh", "未分类")
        atom["impact_labels"] = [taxonomy["impact_signals"].get(value, value) for value in atom.get("impact_signals", [])]
        enriched.append(atom)
    return enriched


def unique_reviews(atoms: Iterable[dict[str, Any]], polarity: str | None = None) -> set[str]:
    return {atom["review_id"] for atom in atoms if polarity is None or atom["polarity"] == polarity}


def product_rows(reviews: list[dict[str, Any]], atoms: list[dict[str, Any]]) -> list[dict[str, Any]]:
    reviews_by_product: dict[str, list[dict[str, Any]]] = defaultdict(list)
    atoms_by_product: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for review in reviews:
        reviews_by_product[review["product_id"]].append(review)
    for atom in atoms:
        atoms_by_product[atom["product_id"]].append(atom)
    result = []
    for product_id, values in reviews_by_product.items():
        product_atoms = atoms_by_product[product_id]
        images = next((row["product_images"] for row in values if row["product_images"]), [])
        result.append({
            "product_id": product_id,
            "entity_role": Counter(row["entity_role"] for row in values).most_common(1)[0][0],
            "brand": Counter(row["brand"] for row in values).most_common(1)[0][0],
            "title": Counter(row["product_name"] for row in values).most_common(1)[0][0],
            "platforms": sorted({row["platform"] for row in values}),
            "markets": sorted({row["market"] for row in values}),
            "languages": sorted({row["language"] for row in values}),
            "review_count": len(values),
            "positive_review_count": len(unique_reviews(product_atoms, "positive")),
            "negative_review_count": len(unique_reviews(product_atoms, "negative")),
            "image_review_count": sum(bool(row["review_images"]) for row in values),
            "image": images[0] if images else None,
            "temporary_product_id": any(row["temporary_product_id"] for row in values),
        })
    return sorted(result, key=lambda row: (row["entity_role"] != "TARGET", -row["review_count"], row["product_id"]))


def classify_opportunity(row: dict[str, Any], target_total: int, peer_total: int, product_total: int) -> str:
    if row["review_count"] < 3:
        return "EMERGING_SIGNAL"
    positive, negative = row["positive_review_count"], row["negative_review_count"]
    if positive >= 2 and negative >= 2 and min(positive, negative) / max(1, row["review_count"]) >= 0.25:
        return "POLARIZING"
    # A target-only dataset can show priorities and gaps inside that product, but
    # cannot support a relative differentiator claim without peer evidence.
    if target_total and peer_total:
        target_pos = row["target_positive_review_count"] / target_total
        target_neg = row["target_negative_review_count"] / target_total
        peer_pos = row["peer_positive_review_count"] / max(1, peer_total)
        peer_neg = row["peer_negative_review_count"] / max(1, peer_total)
        if target_neg >= 0.08 and target_neg > target_pos and peer_neg + 0.03 < target_neg:
            return "CRITICAL_GAP"
        if target_pos >= 0.08 and target_pos - target_neg > peer_pos - peer_neg + 0.06:
            return "HERO_DIFFERENTIATOR"
    if product_total >= 2 and negative >= positive and row["product_count"] >= 2:
        return "WHITESPACE"
    return "TABLE_STAKES" if target_total else "CATEGORY_PRIORITY"


def factor_rows(reviews: list[dict[str, Any]], atoms: list[dict[str, Any]], factors: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    total_reviews = len({row["review_id"] for row in reviews})
    product_total = len({row["product_id"] for row in reviews})
    target_review_ids = {row["review_id"] for row in reviews if row["entity_role"] == "TARGET"}
    peer_review_ids = {row["review_id"] for row in reviews if row["entity_role"] == "PEER"}
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for atom in atoms:
        if atom.get("factor_id"):
            grouped[atom["factor_id"]].append(atom)
    result = []
    for factor_id, values in grouped.items():
        factor = factors.get(factor_id, {"label_zh": factor_id, "family_id": ""})
        reviews_all = unique_reviews(values)
        positive_reviews = unique_reviews(values, "positive")
        negative_reviews = unique_reviews(values, "negative")
        impacted_reviews = {atom["review_id"] for atom in values if atom.get("impact_signals")}
        row = {
            "factor_id": factor_id,
            "factor_label": factor["label_zh"],
            "family_id": factor["family_id"],
            "review_count": len(reviews_all),
            "review_coverage": len(reviews_all) / max(1, total_reviews),
            "positive_review_count": len(positive_reviews),
            "negative_review_count": len(negative_reviews),
            "product_count": len({value["product_id"] for value in values}),
            "brand_count": len({value["brand"] for value in values}),
            "language_count": len({value["language"] for value in values}),
            "behavior_review_count": len(impacted_reviews),
            "behavior_share": len(impacted_reviews) / max(1, len(reviews_all)),
            "target_positive_review_count": len(positive_reviews & target_review_ids),
            "target_negative_review_count": len(negative_reviews & target_review_ids),
            "peer_positive_review_count": len(positive_reviews & peer_review_ids),
            "peer_negative_review_count": len(negative_reviews & peer_review_ids),
            "evidence_atom_ids": [value["atom_id"] for value in sorted(values, key=lambda item: (-int(bool(item.get("impact_signals"))), -float(item.get("confidence", 0))))[:8]],
        }
        row["opportunity_type"] = classify_opportunity(row, len(target_review_ids), len(peer_review_ids), product_total)
        result.append(row)
    return sorted(result, key=lambda row: (row["factor_id"] == "insufficient_detail", -row["review_coverage"], -row["product_count"], -row["behavior_share"], row["factor_label"]))


SURFACES_BY_FAMILY = {
    "need_outcome": ["HERO", "BULLETS", "GALLERY"], "functional_performance": ["HERO", "BULLETS", "GALLERY", "VIDEO"],
    "reliability": ["PDP", "FAQ", "VIDEO"], "sensory_experience": ["GALLERY", "VIDEO", "FAQ"],
    "ease_routine": ["GALLERY", "VIDEO"], "fit_compatibility": ["VARIANT", "GALLERY", "FAQ"],
    "quality_safety": ["PDP", "A_PLUS", "FAQ"], "design_packaging": ["GALLERY", "VIDEO", "PRODUCT"],
    "value": ["BULLETS", "PDP"], "trust_expectation": ["GALLERY", "FAQ"],
    "delivery_service": ["FAQ"], "loyalty_comparison": ["A_PLUS", "GALLERY"], "insufficient_detail": ["PDP"],
}


OPPORTUNITY_LABELS = {
    "TABLE_STAKES": "基础门槛", "CATEGORY_PRIORITY": "品类优先因子", "HERO_DIFFERENTIATOR": "英雄差异点",
    "CRITICAL_GAP": "目标产品关键短板", "WHITESPACE": "品类空白机会", "POLARIZING": "分化/预期敏感", "EMERGING_SIGNAL": "新兴信号",
}


def opportunity_cards(rows: list[dict[str, Any]], taxonomy: dict[str, Any]) -> list[dict[str, Any]]:
    family_labels = {row["id"]: row["label_zh"] for row in taxonomy["families"]}
    result = []
    for row in (item for item in rows if item["factor_id"] != "insufficient_detail"):
        if len(result) >= 16:
            break
        kind = row["opportunity_type"]
        if kind == "HERO_DIFFERENTIATOR":
            direction = "优先验证目标产品的真实优势；证据充分后，把差异点前置到头图与核心卖点。"
        elif kind == "CRITICAL_GAP":
            direction = "先修复产品或预期管理，再决定是否在商品页主动解释边界。"
        elif kind == "WHITESPACE":
            direction = "把跨产品反复出现的痛点转成研发、包装或证据型内容机会。"
        elif kind == "POLARIZING":
            direction = "明确适用人群、使用方式和不适用边界，降低错误预期。"
        elif kind == "EMERGING_SIGNAL":
            direction = "保留为研究假设，补充样本后再进入前台页面。"
        else:
            direction = "作为购买决策基础信息，用简洁证据回答用户最常见的确认问题。"
        result.append({
            **row,
            "opportunity_label": OPPORTUNITY_LABELS[kind],
            "family_label": family_labels.get(row["family_id"], row["family_id"]),
            "decision_question": f"购买前，用户会确认“{row['factor_label']}”是否符合自己的预期吗？",
            "message_direction": direction,
            "page_surfaces": SURFACES_BY_FAMILY.get(row["family_id"], ["PDP"]),
            "proof_suggestion": "优先使用可复核的产品规格、对比测试、使用步骤、场景图或真实演示；不要把评论体验直接改写成产品事实。",
            "fact_required": "需要目标产品事实、测试或合规材料确认可兑现范围。",
            "claim_status": "VOC_ONLY_DRAFT",
        })
    return result


def keyword_rows(atoms: list[dict[str, Any]], polarity: str, kind: str, limit: int = 60) -> list[dict[str, Any]]:
    key = "keywords_original" if kind == "original" else "keywords_canonical_zh"
    grouped: dict[str, dict[str, set[str]]] = defaultdict(lambda: {"reviews": set(), "atoms": set(), "products": set()})
    for atom in atoms:
        if atom["polarity"] != polarity:
            continue
        for keyword in atom.get(key, []):
            normalized = re.sub(r"\s+", " ", text_value(keyword).casefold()).strip(" ,.;，。；")
            if len(normalized) < 2 or normalized in GENERIC_KEYWORDS:
                continue
            grouped[normalized]["reviews"].add(atom["review_id"])
            grouped[normalized]["atoms"].add(atom["atom_id"])
            grouped[normalized]["products"].add(atom["product_id"])
    rows = [{"keyword": keyword, "review_count": len(value["reviews"]), "atom_count": len(value["atoms"]), "product_count": len(value["products"])} for keyword, value in grouped.items()]
    return sorted(rows, key=lambda row: (-row["review_count"], -row["atom_count"], row["keyword"]))[:limit]


def keyword_cloud(atoms: list[dict[str, Any]]) -> dict[str, Any]:
    products = sorted({atom["product_id"] for atom in atoms})
    def bundle(values: list[dict[str, Any]]) -> dict[str, Any]:
        return {kind: {polarity: keyword_rows(values, polarity, kind) for polarity in ("positive", "negative")} for kind in ("original", "canonical")}
    return {"method": "仅聚合已确认标注中的原语言关键词和中文规范词；字号按去重评论覆盖数缩放。", "global": bundle(atoms), "by_product": {product: bundle([atom for atom in atoms if atom["product_id"] == product]) for product in products}}


def factor_matrix_rows(products: list[dict[str, Any]], factor_summary: list[dict[str, Any]], atoms: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for product in products:
        product_atoms = [atom for atom in atoms if atom["product_id"] == product["product_id"]]
        product_reviews = {atom["review_id"] for atom in product_atoms}
        for factor in factor_summary:
            values = [atom for atom in product_atoms if atom["factor_id"] == factor["factor_id"]]
            positive = len(unique_reviews(values, "positive"))
            negative = len(unique_reviews(values, "negative"))
            rows.append({"product_id": product["product_id"], "entity_role": product["entity_role"], "brand": product["brand"], "factor_id": factor["factor_id"], "factor_label": factor["factor_label"], "review_count": len(unique_reviews(values)), "positive_review_count": positive, "negative_review_count": negative, "net_signal": (positive - negative) / max(1, len(product_reviews))})
    return rows


def build_report_data(reviews: list[dict[str, Any]], atoms: list[dict[str, Any]], audit: dict[str, Any], taxonomy: dict[str, Any], catalog: dict[str, Any], factors: dict[str, dict[str, Any]], catalog_confirmed: bool, annotations_supplied: bool) -> dict[str, Any]:
    confirmed = sum(atom["review_status"] in CONFIRMED_STATUSES for atom in atoms)
    unresolved = sum(atom["review_status"] not in CONFIRMED_STATUSES for atom in atoms)
    translation_unresolved = sum(atom["language"] != "zh" and (atom["translation_status"] not in CONFIRMED_TRANSLATIONS or not atom.get("evidence_zh")) for atom in atoms)
    temporary = sum(review["temporary_product_id"] for review in reviews)
    homogeneity = audit.get("category_homogeneity")
    if not reviews or not atoms or homogeneity == "MIXED":
        status = "FAIL"
    elif annotations_supplied and unresolved == 0 and translation_unresolved == 0 and temporary == 0 and homogeneity == "CONFIRMED" and catalog_confirmed:
        status = "PASS"
    else:
        status = "DEGRADED"
    products = product_rows(reviews, atoms)
    factors_summary = factor_rows(reviews, atoms, factors)
    matrix = factor_matrix_rows(products, factors_summary, atoms)
    categories = audit.get("category_values", [])
    summary = [f"本次样本覆盖 {len(products)} 个同品类产品、{len(reviews)} 条去重评论、{len(atoms)} 条原子声音和 {len(audit.get('language_counts', {}))} 种语言。"]
    if audit.get("demo_mode"):
        summary.insert(0, "这是合成演示数据：流程门槛通过只说明报告能力可运行，不可用于真实业务决策。")
    if factors_summary:
        top = factors_summary[0]
        summary.append(f"提及最广的购买决策因子是“{top['factor_label']}”，覆盖 {top['review_count']} 条评论和 {top['product_count']} 个产品。")
    gaps = [row for row in factors_summary if row["opportunity_type"] == "CRITICAL_GAP"]
    heroes = [row for row in factors_summary if row["opportunity_type"] == "HERO_DIFFERENTIATOR"]
    if heroes:
        summary.append(f"目标产品当前最值得验证的差异点是“{heroes[0]['factor_label']}”。")
    if gaps:
        summary.append(f"目标产品优先修复或管理预期的因子是“{gaps[0]['factor_label']}”。")
    if status != "PASS":
        summary.append(f"当前仍有 {unresolved} 条未确认原子和 {translation_unresolved} 条未确认翻译；正式决策前需完成语义复核。")
    return {
        "schema_version": "ecommerce-review-purchase-drivers-v1",
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "demo_mode": bool(audit.get("demo_mode")),
        "category": catalog.get("category") or (categories[0] if len(categories) == 1 else "未确认品类"),
        "quality": {"status": status, "catalog_confirmed": catalog_confirmed, "annotations_supplied": annotations_supplied, "confirmed_atom_count": confirmed, "unresolved_atom_count": unresolved, "translation_unresolved_count": translation_unresolved, "temporary_product_review_count": temporary, "category_homogeneity": homogeneity, "image_failure_count": audit.get("image_audit", {}).get("failure_count", 0)},
        "totals": {"review_count": len(reviews), "atom_count": len(atoms), "product_count": len(products), "brand_count": len({review["brand"] for review in reviews}), "target_product_count": len({review["product_id"] for review in reviews if review["entity_role"] == "TARGET"}), "positive_review_count": len(unique_reviews(atoms, "positive")), "negative_review_count": len(unique_reviews(atoms, "negative")), "review_with_images_count": sum(bool(review["review_images"]) for review in reviews), "duplicate_count": audit.get("duplicate_count", 0), "cross_product_repeated_text_row_count": audit.get("cross_product_repeated_text_row_count", 0)},
        "summary_lines": summary,
        "products": products,
        "factor_summary": factors_summary,
        "factor_matrix": matrix,
        "opportunities": opportunity_cards(factors_summary, taxonomy),
        "keyword_cloud": keyword_cloud(atoms),
        "families": taxonomy["families"],
        "impact_signal_labels": taxonomy["impact_signals"],
        "page_surface_labels": taxonomy["page_surfaces"],
        "reviews": reviews,
        "atoms": atoms,
        "audit": audit,
        "definitions": {"review_coverage": "提及该因子的去重评论数 ÷ 当前筛选的去重评论数。", "product_breadth": "提及该因子的产品数；不代表市场渗透率。", "opportunity_type": "由样本量、产品广度、正负覆盖和目标/竞品差异的透明规则得到，不是黑盒分数。", "claim_gate": "VOC 证明用户关注点；公开宣称仍需产品事实、测试和合规证据。"},
    }


def json_for_html(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":")).replace("<", "\\u003c").replace(">", "\\u003e").replace("&", "\\u0026").replace("\u2028", "\\u2028").replace("\u2029", "\\u2029")


def command_build(args: argparse.Namespace) -> int:
    run_dir = Path(args.run_dir).resolve()
    taxonomy = read_json(TAXONOMY_PATH)
    reviews = read_jsonl(run_dir / "normalized-reviews.jsonl")
    source_atoms = read_jsonl(run_dir / "evidence-atoms.jsonl")
    audit = read_json(run_dir / "import-audit.json")
    catalog_path = Path(args.factor_catalog).resolve() if args.factor_catalog else (run_dir / "factor-catalog.json" if (run_dir / "factor-catalog.json").is_file() else run_dir / "factor-catalog-draft.json")
    catalog = read_json(catalog_path)
    factors, catalog_confirmed = validate_catalog(catalog, taxonomy)
    annotation_path = Path(args.annotations).resolve() if args.annotations else None
    annotations_supplied = bool(annotation_path and annotation_path.is_file())
    annotations = read_jsonl(annotation_path) if annotations_supplied and annotation_path else []
    annotations_by_id = validate_annotations(annotations, source_atoms, factors, taxonomy) if annotations_supplied else {}
    atoms = enrich_atoms(source_atoms, annotations_by_id, factors, taxonomy)
    report = build_report_data(reviews, atoms, audit, taxonomy, catalog, factors, catalog_confirmed, annotations_supplied)
    write_csv(run_dir / "evidence-atoms.csv", atoms, ["atom_id", "review_id", "product_id", "entity_role", "brand", "platform", "market", "language", "polarity", "family_id", "family_label", "factor_id", "factor_label", "evidence_original", "evidence_zh", "keywords_original", "keywords_canonical_zh", "contexts", "impact_signals", "confidence", "translation_status", "review_status"])
    write_csv(run_dir / "decision-factor-matrix.csv", report["factor_matrix"], ["product_id", "entity_role", "brand", "factor_id", "factor_label", "review_count", "positive_review_count", "negative_review_count", "net_signal"])
    write_csv(run_dir / "pdp-opportunities.csv", report["opportunities"], ["factor_id", "factor_label", "family_label", "opportunity_type", "opportunity_label", "decision_question", "review_count", "review_coverage", "product_count", "positive_review_count", "negative_review_count", "target_positive_review_count", "target_negative_review_count", "peer_positive_review_count", "peer_negative_review_count", "message_direction", "page_surfaces", "proof_suggestion", "fact_required", "claim_status"])
    write_json(run_dir / "analysis.json", report)
    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    if "__REPORT_DATA__" not in template:
        raise RuntimeError("HTML 模板缺少 __REPORT_DATA__")
    html = template.replace("__REPORT_DATA__", json_for_html(report))
    output = Path(args.output).resolve() if args.output else run_dir / "review-purchase-drivers.html"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(html, encoding="utf-8")
    summary = {"status": report["quality"]["status"], "report": str(output), "reviews": report["totals"]["review_count"], "products": report["totals"]["product_count"], "atoms": report["totals"]["atom_count"], "unresolved_atoms": report["quality"]["unresolved_atom_count"], "unresolved_translations": report["quality"]["translation_unresolved_count"]}
    write_json(run_dir / "run-summary.json", summary)
    print(json.dumps(summary, ensure_ascii=False))
    return 0


def command_verify(args: argparse.Namespace) -> int:
    run_dir = Path(args.run_dir).resolve()
    required = ["normalized-reviews.csv", "normalized-reviews.jsonl", "evidence-atoms.jsonl", "evidence-atoms.csv", "factor-catalog-draft.json", "import-audit.json", "analysis.json", "decision-factor-matrix.csv", "pdp-opportunities.csv", "run-summary.json", "review-purchase-drivers.html"]
    missing = [name for name in required if not (run_dir / name).is_file()]
    errors = ["缺少文件：" + ", ".join(missing)] if missing else []
    report_path = run_dir / "review-purchase-drivers.html"
    if report_path.is_file():
        html = report_path.read_text(encoding="utf-8")
        if "__REPORT_DATA__" in html:
            errors.append("HTML 数据占位符未替换")
        for marker in ('id="decision-map"', 'id="factor-matrix"', 'id="opportunity-list"', 'id="positive-word-cloud"', 'id="negative-word-cloud"', 'id="evidence-list"'):
            if marker not in html:
                errors.append(f"HTML 缺少关键区域：{marker}")
    analysis = read_json(run_dir / "analysis.json") if (run_dir / "analysis.json").is_file() else {}
    status = analysis.get("quality", {}).get("status", "FAIL")
    result = {"valid": not errors, "report_status": status, "errors": errors, "run_dir": str(run_dir)}
    print(json.dumps(result, ensure_ascii=False))
    return 1 if errors else 0 if status == "PASS" else 2


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="同品类多产品电商评论购买决策因子分析")
    commands = parser.add_subparsers(dest="command", required=True)
    prepare = commands.add_parser("prepare", help="归一化评论并生成因子发现与语义标注队列")
    prepare.add_argument("--input", action="append", default=[], help="混合评论文件；可重复")
    prepare.add_argument("--target", action="append", default=[], help="目标产品评论文件；可重复")
    prepare.add_argument("--peer", action="append", default=[], help="竞品评论文件；可重复")
    prepare.add_argument("--positive", action="append", default=[], help="明确为好评的文件；可重复")
    prepare.add_argument("--negative", action="append", default=[], help="明确为差评的文件；可重复")
    prepare.add_argument("--target-product-id", default="", help="把匹配的产品 ID 标记为目标产品")
    prepare.add_argument("--category", default="", help="本次唯一可比较品类；正式报告必须确认")
    prepare.add_argument("--output-dir", required=True)
    prepare.add_argument("--download-images", action="store_true", help="安全下载并缓存公网评论/产品图片")
    prepare.add_argument("--demo", action="store_true", help="标记为合成/演示数据；报告不得描述为可用于真实业务决策")
    prepare.set_defaults(func=command_prepare)
    build = commands.add_parser("build", help="合并已确认因子目录和标注，生成交互式 HTML")
    build.add_argument("--run-dir", required=True)
    build.add_argument("--factor-catalog", help="已确认的运行级因子目录；默认查找 factor-catalog.json")
    build.add_argument("--annotations", help="AI/人工确认的 annotations.jsonl")
    build.add_argument("--output", help="HTML 输出路径")
    build.set_defaults(func=command_build)
    verify = commands.add_parser("verify", help="验证运行包和正式交付状态")
    verify.add_argument("--run-dir", required=True)
    verify.set_defaults(func=command_verify)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        return int(args.func(args))
    except Exception as exc:
        print(json.dumps({"status": "ERROR", "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
